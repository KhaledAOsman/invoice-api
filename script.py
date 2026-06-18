
import os

os.makedirs("pdf-to-excel/templates", exist_ok=True)

# ============================================================
# app.py
# ============================================================
app_py = '''import os
import uuid
import tempfile
from flask import Flask, request, send_file, jsonify, render_template

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50MB

UPLOAD_FOLDER = tempfile.gettempdir()

def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() == "pdf"

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/convert", methods=["POST"])
def convert():
    if "pdf" not in request.files:
        return jsonify({"error": "لم يتم إرسال ملف"}), 400

    file = request.files["pdf"]

    if not file.filename or not allowed_file(file.filename):
        return jsonify({"error": "يجب أن يكون الملف بصيغة PDF"}), 400

    uid = str(uuid.uuid4())[:8]
    pdf_path   = os.path.join(UPLOAD_FOLDER, f"{uid}_input.pdf")
    excel_path = os.path.join(UPLOAD_FOLDER, f"{uid}_output.xlsx")

    file.save(pdf_path)

    try:
        from invoice_parser import extract_all_invoices, to_excel_merged
        all_invoices = extract_all_invoices(pdf_path)

        if not all_invoices:
            return jsonify({"error": "لم يتم العثور على فواتير في الملف"}), 422

        to_excel_merged(all_invoices, excel_path)

        return send_file(
            excel_path,
            as_attachment=True,
            download_name="invoices_output.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        return jsonify({"error": f"خطأ أثناء المعالجة: {str(e)}"}), 500
    finally:
        if os.path.exists(pdf_path):
            os.remove(pdf_path)

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
'''

with open("pdf-to-excel/app.py", "w", encoding="utf-8") as f:
    f.write(app_py)

# ============================================================
# requirements.txt
# ============================================================
requirements = """flask==3.0.0
gunicorn==21.2.0
pdfplumber==0.11.0
pandas==2.2.2
openpyxl==3.1.4
arabic-reshaper==3.0.0
python-bidi==0.6.6
"""

with open("pdf-to-excel/requirements.txt", "w") as f:
    f.write(requirements)

# ============================================================
# render.yaml
# ============================================================
render_yaml = """services:
  - type: web
    name: pdf-to-excel
    runtime: python
    buildCommand: pip install -r requirements.txt
    startCommand: gunicorn app:app
    envVars:
      - key: PYTHON_VERSION
        value: 3.11.0
"""

with open("pdf-to-excel/render.yaml", "w") as f:
    f.write(render_yaml)

# ============================================================
# invoice_parser.py — كود المستخدم الكامل
# ============================================================
invoice_parser = '''import re
import json
import unicodedata
import arabic_reshaper
from bidi.algorithm import get_display
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

# ============================================================
# HELPERS
# ============================================================

def fix_arabic(text):
    if not text:
        return text
    return unicodedata.normalize("NFKC", get_display(arabic_reshaper.reshape(text)))


def normalize_text(text):
    if not text:
        return ""
    text = text.replace("\\u200f", " ").replace("\\u200e", " ").replace("\\xa0", " ")
    text = re.sub(r"[ \\t]+", " ", text)
    return re.sub(r"\\n+", "\\n", text).strip()


def nfkc(text):
    return unicodedata.normalize("NFKC", text or "")


def parse_sar(text):
    m = re.search(r"SAR\\s*([\\d,]+(?:\\.\\d+)?)", str(text or ""))
    return f"SAR {m.group(1)}" if m else None


def to_float(s):
    m = re.search(r"[\\d,]+(?:\\.\\d+)?", str(s or ""))
    return float(m.group().replace(",", "")) if m else None


STORE_PHONES = {"+966533343208"}

# ============================================================
# PATTERNS
# ============================================================

PRODUCT_LINE = re.compile(
    r"^SAR\\s*([\\d,]+(?:\\.\\d+)?)\\s+SAR\\s*([\\d,]+(?:\\.\\d+)?)\\s+(\\d+)\\s+(.+)$"
)
VAT_LINE = re.compile(
    r"^SAR\\s*([\\d,]+(?:\\.\\d+)?)\\s*:\\s*\\(%15\\)\\s*\\ufebc\\ufea4\\ufef3\\ufed0\\ufea3\\ufeb7(.*)$"
)
PRICE_INCL = re.compile(
    r"^SAR\\s*([\\d,]+(?:\\.\\d+)?)\\s*:\\s*\\ufebc\\ufea4\\ufef3\\ufed0\\ufea3\\ufeb7 \\ufeb7\\ufee4\\ufea7\\ufeb7 \\ufef2\\ufec0\\ufea4\\ufeb7"
)
BEFORE_DISCOUNT = re.compile(
    r"^SAR\\s*([\\d,]+(?:\\.\\d+)?)\\s*:\\s*\\ufed0\\ufee4\\ufef3\\ufeb3\\ufee3\\ufed9\\ufeb7 \\ufeb7\\ufeb7\\ufea3 \\ufef2\\ufec0\\ufea4\\ufeb7"
)
OPTION_TYPES = ["\\ufebc\\ufef3\\ufee4\\ufee9\\ufeb7", "\\ufee6\\ufec2\\ufee0\\ufea7", "\\ufee4\\ufea7\\ufef2\\ufee4", "\\ufee6\\ufee0\\ufee0\\ufea3", "\\ufee4\\ufea4\\ufea4\\ufea7"]

# ============================================================
# CLASSIFIERS
# ============================================================

def is_invoice_start(l):
    return l.strip() == "\\ufead\\ufef2\\ufeab\\ufea7\\ufea1\\ufea1"


def is_invoice_end(l):
    return "\\ufea7\\ufef2\\ufe97\\ufeb7" in l and "\\ufef2\\ufea4\\ufeb3\\ufee3\\ufea3\\ufeb7" in l


def is_subtotal(l):
    return ("\\ufef3\\ufec4\\ufea7\\ufee4\\ufea4\\ufef3\\ufea7" in l or "\\ufef3\\ufec4\\ufea3\\ufee4\\ufea4\\ufef3\\ufea7" in l) and ("\\ufef3\\ufec0\\ufef2\\ufef5\\ufea3\\ufeb7" in l or "\\ufef3\\ufec0\\ufef2\\ufec0\\ufeb7" in l)


def is_shipping(l):
    return "\\ufee6\\ufea4\\ufea4\\ufeb7 \\ufebc\\ufef5\\ufeac\\ufea7\\ufee3\\ufea7" in l


def is_ototal(l):
    return "\\ufeb7\\ufeb4\\ufeb7\\ufeb4\\ufb04" in l and ("\\ufef3\\ufec4\\ufea7\\ufee4\\ufea4\\ufef3\\ufef2" in l or "\\ufef3\\ufec4\\ufea3\\ufee4\\ufea4\\ufef3\\ufef2" in l)


def is_vat_tot(l):
    return "%" in l and "15" in l and "\\ufebc\\ufea4\\ufef3\\ufed0\\ufea3\\ufeb7" in l and VAT_LINE.match(l) is None and PRICE_INCL.match(l) is None


def is_opt_label(l):
    return "\\ufea4\\ufea4\\ufee3\\ufea7\\ufeb7 \\ufee3\\ufea7\\ufef2\\ufea7\\ufef3\\ufea7\\ufeaf" in l


def is_tbl_hdr(l):
    return "\\ufec0\\ufee0\\ufee4\\ufea4\\ufee4\\ufea7" in l and "\\ufef2\\ufea4\\ufea4\\ufeb7" in l and "\\ufebc\\ufef3\\ufee4\\ufee9\\ufeb7" in l


def parse_option_line(line):
    for opt_type in OPTION_TYPES:
        if line.endswith(opt_type):
            return fix_arabic(opt_type), fix_arabic(line[:-len(opt_type)].strip())
    parts = line.rsplit(None, 1)
    if len(parts) == 2:
        return fix_arabic(parts[1]), fix_arabic(parts[0])
    return None, fix_arabic(line)

# ============================================================
# PAYMENT METHOD DETECTOR
# ============================================================

PAYMENT_METHODS = [
    ("تابي", [
        "TABBY", "تابي", "تابى", "يبات", "بيات", "تاي", "\\ufef3\\ufec4\\ufea7\\ufee3", "\\ufef3\\ufec4\\ufea7\\ufee3\\ufeb7",
    ]),
    ("تمارا", [
        "TAMARA", "تمارا", "ارامت", "\\ufea7\\ufef2\\ufea7\\ufee4\\ufee3", "\\ufea7\\ufef2\\ufea7\\ufee4\\ufee3",
    ]),
    ("مدى", [
        "MADA", "مدى", "\\ufef3\\ufed0\\ufee4", "\\ufef3\\ufed0\\ufee4\\ufee4",
    ]),
    ("البطاقة الائتمانية", [
        "CREDIT CARD", "CREDITCARD", "\\ufebc\\ufef3\\ufee0\\ufea7\\ufee4\\ufee3\\ufee3\\ufea7\\ufea7",
        "\\ufebc\\ufea4\\ufea7\\ufeb4\\ufea4\\ufeb7",
    ]),
    ("Apple Pay",  ["APPLE PAY", "APPLEPAY"]),
    ("STC Pay",    ["STC PAY", "STCPAY"]),
    ("Visa",       ["VISA"]),
    ("Mastercard", ["MASTERCARD", "MASTER CARD"]),
    ("أرامكس",     ["ARAMEX", "\\ufebc\\ufee9\\ufee4\\ufea7\\ufef2", "\\ufebc\\ufee9\\ufee4\\ufea7\\ufef2\\ufef2"]),
    ("مرابحة",     ["MURABAHA", "\\ufebc\\ufea4\\ufea7\\ufec4\\ufea7\\ufef2\\ufee4", "\\ufebc\\ufea4\\ufea7\\ufec4\\ufea7\\ufef2\\ufee4"]),
]


def detect_payment_method(line, next_line=""):
    combined      = line + " " + next_line
    combined_nfkc = nfkc(combined).upper()
    for method_name, keywords in PAYMENT_METHODS:
        for kw in keywords:
            if nfkc(kw).upper() in combined_nfkc:
                return method_name
    return None

# ============================================================
# SPLIT INTO INVOICE BLOCKS
# ============================================================

def split_invoices(lines):
    invoices, current, in_invoice = [], [], False
    for line in lines:
        if is_invoice_start(line):
            if current:
                invoices.append(current)
            current = [line]
            in_invoice = True
            continue
        if in_invoice:
            current.append(line)
            if is_invoice_end(line):
                invoices.append(current)
                current = []
                in_invoice = False
    if current:
        invoices.append(current)
    return invoices

# ============================================================
# PRODUCTS + TOTALS PARSER
# ============================================================

def parse_products_and_totals(lines):
    products, totals = [], {}
    tbl_start = None

    for i, l in enumerate(lines):
        if is_tbl_hdr(l):
            tbl_start = i + 1
            break

    if tbl_start is None:
        for i, l in enumerate(lines):
            if PRODUCT_LINE.match(l):
                tbl_start = i
                break

    if tbl_start is None:
        return products, totals

    i = tbl_start
    current = None

    def save():
        nonlocal current
        if current and current.get("qty") is not None:
            products.append(dict(current))

    while i < len(lines):
        l = lines[i].strip()

        if is_subtotal(l):
            save(); current = None
            v = parse_sar(l)
            totals["subtotal"] = v
            totals["subtotal_value"] = to_float(v)
            i += 1; continue

        if is_shipping(l):
            if "\\ufea7\\ufea7\\ufea7\\ufea7\\ufea4\\ufee4" in l or "\\ufea7\\ufea7\\ufea7\\ufea7\\ufea3\\ufee4" in l:
                totals["shipping"] = "مجانًا"
                totals["shipping_value"] = 0.0
            else:
                v = parse_sar(l)
                totals["shipping"] = v
                totals["shipping_value"] = to_float(v)
            i += 1; continue

        if is_vat_tot(l) and current is None:
            v = parse_sar(l)
            totals["vat_15"] = v
            totals["vat_15_value"] = to_float(v)
            i += 1; continue

        if is_ototal(l):
            v = parse_sar(l)
            totals["order_total"] = v
            totals["order_total_value"] = to_float(v)
            i += 1; continue

        m = PRODUCT_LINE.match(l)
        if m:
            save()
            lt = f"SAR {m.group(1)}"
            up = f"SAR {m.group(2)}"
            current = {
                "name": fix_arabic(m.group(4).strip()),
                "description": "",
                "qty": int(m.group(3)),
                "unit_price": up,
                "unit_price_value": to_float(up),
                "line_total": lt,
                "line_total_value": to_float(lt),
                "vat": None, "vat_value": None,
                "price_incl_vat": None, "price_incl_vat_value": None,
                "option_type": None, "option_value": None,
                "barcode": None, "discount_price": None,
            }
            i += 1; continue

        m = VAT_LINE.match(l)
        if m and current:
            v = f"SAR {m.group(1)}"
            current["vat"] = v
            current["vat_value"] = to_float(v)
            desc = m.group(2).strip()
            if desc:
                current["description"] = fix_arabic(desc)
            i += 1; continue

        m = PRICE_INCL.match(l)
        if m and current:
            v = f"SAR {m.group(1)}"
            current["price_incl_vat"] = v
            current["price_incl_vat_value"] = to_float(v)
            i += 1; continue

        m = BEFORE_DISCOUNT.match(l)
        if m and current:
            current["discount_price"] = f"SAR {m.group(1)}"
            i += 1; continue

        if is_opt_label(l) and current:
            i += 1
            if i < len(lines):
                opt_type, opt_val = parse_option_line(lines[i].strip())
                current["option_type"] = opt_type
                current["option_value"] = opt_val
                i += 1
            continue

        if current and re.match(r"^\\d{2,6}-\\d+$", l):
            current["barcode"] = l
            i += 1; continue

        i += 1

    save()
    return products, totals

# ============================================================
# HEADER PARSER
# ============================================================

def parse_header(lines, all_words):
    r = {}

    for w in all_words:
        t = w.get("text", "")
        if re.match(r"^311\\d{12}$", t):
            r["vat_number"] = t
        if re.match(r"^\\+966\\d{9}$", t) and t not in STORE_PHONES:
            r.setdefault("customer_phone", t)

    for i, line in enumerate(lines):
        next_line = lines[i + 1] if i + 1 < len(lines) else ""

        m = re.match(r"^(\\d{6,9})\\s*:\\s*\\ufeb7\\ufeb4\\ufeb7\\ufeb4\\ufb04", line)
        if not m:
            m = re.match(r"^(\\d{6,9})\\s*:", line)
        if m and "order_number" not in r:
            r["order_number"] = m.group(1)

        if "vat_number" not in r:
            m = re.search(r"(311\\d{12})", line)
            if m:
                r["vat_number"] = m.group(1)

        m = re.match(
            r"((?:Saturday|Sunday|Monday|Tuesday|Wednesday|Thursday|Friday)"
            r"\\s+\\d{1,2}\\s+\\w+\\s+\\d{4})\\s*\\|",
            line,
        )
        if m and "order_date" not in r:
            dp = m.group(1)
            tm = re.search(r"(\\d{2}:\\d{2})", line)
            ap = re.search(r"\\b(AM|PM)\\b", line)
            if tm:
                ampm = ap.group(1) if ap else ""
                if not ampm:
                    ap2 = re.search(r"\\b(AM|PM)\\b", next_line)
                    ampm = ap2.group(1) if ap2 else ""
                r["order_date"] = f"{dp} | {tm.group(1)} {ampm}".strip()
            else:
                t2 = re.search(r"(\\d{2}:\\d{2}\\s*(?:AM|PM))", next_line)
                if t2:
                    r["order_date"] = f"{dp} | {t2.group(1).strip()}"

        if "\\ufec0\\ufeb4\\ufeb4\\ufea7" in line and "payment_amount" not in r:
            m = re.search(r"SAR\\s*([\\d,]+(?:\\.\\d+)?)", line)
            if m:
                r["payment_amount"] = f"SAR {m.group(1)}"

        line_nfkc = nfkc(line)
        if "\\u0639\\u0641\\u062f\\u0644\\u0627" in line_nfkc and "\\u0629\\u0642\\u064a\\u0631\\u0637" in line_nfkc:
            if "payment_method" not in r:
                pm = detect_payment_method(line, next_line)
                if pm:
                    r["payment_method"] = pm
                else:
                    for look in range(1, 4):
                        idx = i + look
                        if idx < len(lines):
                            pm = detect_payment_method(lines[idx])
                            if pm:
                                r["payment_method"] = pm
                                break

    h = next((i for i, l in enumerate(lines) if "\\ufead\\ufef2\\ufeb7\\ufed0\\ufee4" in l and l.count("\\ufead\\ufef2\\ufeb7\\ufed0\\ufee4") >= 2), None)
    if h is not None and h + 1 < len(lines):
        name_line = lines[h + 1]
        for kw in ["\\ufef2\\ufea4\\ufeb3\\ufee3\\ufea3\\ufeb7", "\\ufebc\\ufef3\\ufec0\\ufea7\\ufeb3", "\\ufead\\ufee0\\ufec0\\ufea3\\ufeb7", "\\ufee6\\ufef3\\ufeab\\ufef2\\ufee9\\ufea7\\ufef3\\ufef3\\ufef2"]:
            idx = name_line.find(kw)
            if idx != -1:
                name_line = name_line[:idx].strip()
        r["customer_name"] = fix_arabic(name_line.strip())
        if h + 2 < len(lines):
            r["customer_country"] = fix_arabic(lines[h + 2].split()[0])
        if h + 3 < len(lines):
            r["customer_city"] = fix_arabic(lines[h + 3].split()[0])

    for line in lines:
        for e in re.findall(r"\\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\\.[A-Za-z]{2,}\\b", line):
            if "vagueoud" not in e:
                r["customer_email"] = e
                break
        if "customer_email" in r:
            break

    return r

# ============================================================
# PARSE SINGLE INVOICE
# ============================================================

def parse_single_invoice(inv_lines, all_words):
    invoice = parse_header(inv_lines, all_words)
    products, totals = parse_products_and_totals(inv_lines)
    return {
        "invoice": invoice,
        "products": products,
        "totals": totals,
        "summary": {
            "total_products": len(products),
            "total_items": sum(p.get("qty") or 0 for p in products),
        },
    }


def extract_all_invoices(pdf_path):
    all_texts, all_words = [], []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            t = normalize_text(page.extract_text() or "")
            all_texts.append(t)
            all_words.extend(page.extract_words() or [])

    all_lines = [l.strip() for l in "\\n".join(all_texts).splitlines() if l.strip()]
    blocks = split_invoices(all_lines)

    results = []
    for idx, block in enumerate(blocks, 1):
        parsed = parse_single_invoice(block, all_words)
        parsed["invoice_index"] = idx
        results.append(parsed)
    return results

# ============================================================
# BUILD FLAT TABLE
# ============================================================

INVOICE_COLS = ["order_number", "order_date", "payment_method", "payment_amount", "shipping"]
PRODUCT_COLS = ["product_index", "product_name", "qty", "unit_price", "option_type", "option_value"]


def build_merged_table(all_invoices):
    rows = []
    for inv in all_invoices:
        h    = inv["invoice"]
        tot  = inv["totals"]
        prods = inv["products"]
        order_num = h.get("order_number", f"INV-{inv[\'invoice_index\']}")
        shipping  = tot.get("shipping") or ""

        inv_data = {
            "order_number":   order_num,
            "order_date":     h.get("order_date", ""),
            "payment_method": h.get("payment_method", ""),
            "payment_amount": h.get("payment_amount", ""),
            "shipping":       shipping,
        }

        if not prods:
            row = dict(inv_data)
            row.update({c: None for c in PRODUCT_COLS})
            rows.append(row)
        else:
            for p_idx, p in enumerate(prods, 1):
                row = dict(inv_data)
                row["product_index"] = p_idx
                row["product_name"]  = p.get("name", "")
                row["qty"]           = p.get("qty")
                row["unit_price"]    = p.get("unit_price", "")
                row["option_type"]   = p.get("option_type", "")
                row["option_value"]  = p.get("option_value", "")
                rows.append(row)

    return pd.DataFrame(rows, columns=INVOICE_COLS + PRODUCT_COLS)

# ============================================================
# EXCEL WITH MERGE CELLS
# ============================================================

def style_border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def to_excel_merged(all_invoices, output_path="invoices_output.xlsx"):
    df = build_merged_table(all_invoices)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="الفواتير", index=False)

    wb = load_workbook(output_path)
    ws = wb["الفواتير"]

    HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
    INV_FILL  = PatternFill("solid", fgColor="D6E4F0")
    INV_FILL2 = PatternFill("solid", fgColor="EBF3FB")
    PROD_FILL = PatternFill("solid", fgColor="FFFFFF")
    ALT_FILL  = PatternFill("solid", fgColor="F5F9FD")

    HDR_FONT   = Font(bold=True, color="FFFFFF", size=11)
    BODY_FONT  = Font(size=10)
    MERGE_FONT = Font(bold=True, size=10)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.fill      = HDR_FILL
        cell.font      = HDR_FONT
        cell.alignment = CENTER
        cell.border    = style_border()
    ws.row_dimensions[1].height = 32

    col_widths = {
        "order_number": 18, "order_date": 28, "payment_method": 22,
        "payment_amount": 18, "shipping": 16, "product_index": 10,
        "product_name": 42, "qty": 8, "unit_price": 16,
        "option_type": 18, "option_value": 22,
    }
    for col_idx, col_name in enumerate(INVOICE_COLS + PRODUCT_COLS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 16)

    order_rows = defaultdict(list)
    for df_idx, val in enumerate(df["order_number"], start=2):
        order_rows[str(val)].append(df_idx)

    inv_col_n  = len(INVOICE_COLS)
    prod_col_n = len(PRODUCT_COLS)

    for inv_i, (order_num, row_list) in enumerate(order_rows.items()):
        inv_fill = INV_FILL if inv_i % 2 == 0 else INV_FILL2
        r_start  = row_list[0]
        r_end    = row_list[-1]

        for col_idx in range(1, inv_col_n + 1):
            col_letter = get_column_letter(col_idx)
            if r_start != r_end:
                ws.merge_cells(f"{col_letter}{r_start}:{col_letter}{r_end}")
            cell = ws[f"{col_letter}{r_start}"]
            cell.fill      = inv_fill
            cell.font      = MERGE_FONT
            cell.alignment = CENTER
            cell.border    = style_border()

        for row_num in row_list:
            fill = PROD_FILL if row_num % 2 == 0 else ALT_FILL
            for col_idx in range(inv_col_n + 1, inv_col_n + prod_col_n + 1):
                cell = ws[f"{get_column_letter(col_idx)}{row_num}"]
                cell.fill      = fill
                cell.font      = BODY_FONT
                cell.alignment = CENTER if col_idx in [inv_col_n + 1, inv_col_n + 3] else LEFT
                cell.border    = style_border()
            ws.row_dimensions[row_num].height = 28

    ws.freeze_panes           = "A2"
    ws.sheet_view.rightToLeft = True
    wb.save(output_path)
    return output_path, df
'''

with open("pdf-to-excel/invoice_parser.py", "w", encoding="utf-8") as f:
    f.write(invoice_parser)

# ============================================================
# templates/index.html — نفس الواجهة السابقة
# ============================================================
html = open("pdf-to-excel-app/templates/index.html", encoding="utf-8").read()
with open("pdf-to-excel/templates/index.html", "w", encoding="utf-8") as f:
    f.write(html)

print("✅ جميع الملفات جاهزة:")
for root, dirs, files in os.walk("pdf-to-excel"):
    level = root.replace("pdf-to-excel", "").count(os.sep)
    indent = "  " * level
    print(f"{indent}📁 {os.path.basename(root)}/")
    for file in files:
        size = os.path.getsize(os.path.join(root, file))
        print(f"{indent}  📄 {file}  ({size:,} bytes)")
