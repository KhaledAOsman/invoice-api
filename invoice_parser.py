import re
import json
import unicodedata
import arabic_reshaper
from bidi.algorithm import get_display
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

# ============================================================
# HELPERS
# ============================================================

def fix_arabic(text):
    if not text:
        return text
    return unicodedata.normalize("NFKC", get_display(arabic_reshaper.reshape(text)))


def normalize_text(text):
    if not text:
        return ""
    text = text.replace("\u200f", " ").replace("\u200e", " ").replace("\xa0", " ")
    text = re.sub(r"[ \t]+", " ", text)
    return re.sub(r"\n+", "\n", text).strip()


def nfkc(text):
    return unicodedata.normalize("NFKC", text or "")


def parse_sar(text):
    m = re.search(r"SAR\s*([\d,]+(?:\.\d+)?)", str(text or ""))
    return f"SAR {m.group(1)}" if m else None


def to_float(s):
    m = re.search(r"[\d,]+(?:\.\d+)?", str(s or ""))
    return float(m.group().replace(",", "")) if m else None


STORE_PHONES = {"+966533343208"}

# ============================================================
# PATTERNS
# ============================================================

PRODUCT_LINE = re.compile(
    r"^SAR\s*([\d,]+(?:\.\d+)?)\s+SAR\s*([\d,]+(?:\.\d+)?)\s+(\d+)\s+(.+)$"
)
VAT_LINE = re.compile(
    r"^SAR\s*([\d,]+(?:\.\d+)?)\s*:\s*\(%15\)\s*\ufebc\ufea4\ufef3\ufed0\ufea3\ufeb7(.*)$"
)
PRICE_INCL = re.compile(
    r"^SAR\s*([\d,]+(?:\.\d+)?)\s*:\s*\ufebc\ufea4\ufef3\ufed0\ufea3\ufeb7 \ufeb7\ufee4\ufea7\ufeb7 \ufef2\ufec0\ufea4\ufeb7"
)
BEFORE_DISCOUNT = re.compile(
    r"^SAR\s*([\d,]+(?:\.\d+)?)\s*:\s*\ufed0\ufee4\ufef3\ufeb3\ufee3\ufed9\ufeb7 \ufeb7\ufeb7\ufea3 \ufef2\ufec0\ufea4\ufeb7"
)
OPTION_TYPES = ["\ufebc\ufef3\ufee4\ufee9\ufeb7", "\ufee6\ufec2\ufee0\ufea7", "\ufee4\ufea7\ufef2\ufee4", "\ufee6\uf
