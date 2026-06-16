import re
import json
import unicodedata
import arabic_reshaper
from bidi.algorithm import get_display
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

# ============================================================
# HELPERS
# ============================================================

def fix_arabic(text):
    if not text:
        return text
    return unicodedata.normalize("NFKC", get_display(arabic_reshaper.reshape(text)))


def normalize_text(text):
    if not text:
        return ""
    text = text.replace("\u200f", " ").replace("\u200e", " ").replace("\xa0", " ")
    text = re.sub(r"[ \t]+", " ", text)
    return re.sub(r"\n+", "\n", text).strip()


def nfkc(text):
    return unicodedata.normalize("NFKC", text or "")


def parse_sar(text):
    m = re.search(r"SAR\s*([\d,]+(?:\.\d+)?)", str(text or ""))
    return f"SAR {m.group(1)}" if m else None


def to_float(s):
    m = re.search(r"[\d,]+(?:\.\d+)?", str(s or ""))
    return float(m.group().replace(",", "")) if m else None


STORE_PHONES = {"+966533343208"}

# ============================================================
# PATTERNS
# ============================================================

PRODUCT_LINE = re.compile(
    r"^SAR\s*([\d,]+(?:\.\d+)?)\s+SAR\s*([\d,]+(?:\.\d+)?)\s+(\d+)\s+(.+)$"
)
VAT_LINE = re.compile(
    r"^SAR\s*([\d,]+(?:\.\d+)?)\s*:\s*\(%15\)\s*ﺔﺒﻳضرﻟا(.*)$"
)
PRICE_INCL = re.compile(
    r"^SAR\s*([\d,]+(?:\.\d+)?)\s*:\s*ﺔﺒﻳضرﻟا ﻞﻣﺎﺷ ﺮﻌﺴﻟا"
)
BEFORE_DISCOUNT = re.compile(
    r"^SAR\s*([\d,]+(?:\.\d+)?)\s*:\s*ﺾﻴﻔﺨﺘﻟا ﻞﺒﻗ ﺮﻌﺴﻟا"
)
OPTION_TYPES = ["ﺔﻴﻤﻜﻟا", "نزﻮﻟا", "ماﺮﺟ", "نﻮﻠﻟا", "مﺠﺤﻟا"]

# ============================================================
# CLASSIFIERS
# ============================================================

def is_invoice_start(l):
    return l.strip() == "ةرﻮﺗﺎﻓ"


def is_invoice_end(l):
    return "اﺮﻜﺷ" in l and "ﺮﺠﺘﻤﻟا" in l


def is_subtotal(l):
    return ("يلﺎﻤﺟﻹا" in l or "ليﺎﻤﺟﻹا" in l) and ("يﻋﺮﻔﻟا" in l or "ﻲﻋﺮﻔﻟا" in l)


def is_shipping(l):
    return "ﻦﺤﺸﻟا ﺔﻔﻠﻜﺗ" in l


def is_ototal(l):
    return "ﺐﻠﻄﻟا" in l and ("يلﺎﻤﺟإ" in l or "ليﺎﻤﺟإ" in l)


def is_vat_tot(l):
    return "%" in l and "15" in l and "ﺔﺒﻳضرﻟا" in l and VAT_LINE.match(l) is None and PRICE_INCL.match(l) is None


def is_opt_label(l):
    return "ﺞﺘﻨﻤﻟا تارﺎﻴﺧ" in l


def is_tbl_hdr(l):
    return "عﻮﻤﺠﻤﻟا" in l and "ﺮﻌﺴﻟا" in l and "ﺔﻴﻤﻜﻟا" in l


def parse_option_line(line):
    for opt_type in OPTION_TYPES:
        if line.endswith(opt_type):
            return fix_arabic(opt_type), fix_arabic(line[:-len(opt_type)].strip())
    parts = line.rsplit(None, 1)
    if len(parts) == 2:
        return fix_arabic(parts[1]), fix_arabic(parts[0])
    return None, fix_arabic(line)

# ============================================================
# PAYMENT METHOD DETECTOR
# ============================================================

PAYMENT_METHODS = [
    ("تابي", [
        "TABBY",
        "تابي",
        "تابى",
        "يبات",
        "بيات",
        "تاي",
        "ﻲﺑﺎﺗ",
        "ﻲﺑﺎﺘﻟا",
    ]),
    ("تمارا", [
        "TAMARA",
        "تمارا",
        "ارامت",
        "اراﻤﺗ",
        "ارﺎﻤﺗ",
    ]),
    ("مدى", [
        "MADA",
        "مدى",
        "ىدم",
        "ىﺪﻣ",
    ]),
    ("البطاقة الائتمانية", [
        "CREDIT CARD",
        "CREDITCARD",
        "ةينامتئلاا",
        "ةقاطبلا",
        "ﺔﻴﻧﺎﻤﺘﺋﻻا",
    ]),
    ("Apple Pay",  ["APPLE PAY", "APPLEPAY"]),
    ("STC Pay",    ["STC PAY", "STCPAY"]),
    ("Visa",       ["VISA"]),
    ("Mastercard", ["MASTERCARD", "MASTER CARD"]),
    ("أرامكس",     ["ARAMEX", "ﺲﻜﻣارأ", "سكمارأ"]),
    ("مرابحة",     ["MURABAHA", "ﺔﺤﺑاﺮﻣ", "ةحباρم", "ةحبارم"]),
]


def detect_payment_method(line, next_line=""):
    combined      = line + " " + next_line
    combined_nfkc = nfkc(combined).upper()
    for method_name, keywords in PAYMENT_METHODS:
        for kw in keywords:
            if nfkc(kw).upper() in combined_nfkc:
                return method_name
    return None

# ============================================================
# SPLIT INTO INVOICE BLOCKS
# ============================================================

def split_invoices(lines):
    invoices, current, in_invoice = [], [], False
    for line in lines:
        if is_invoice_start(line):
            if current:
                invoices.append(current)
            current = [line]
            in_invoice = True
            continue
        if in_invoice:
            current.append(line)
            if is_invoice_end(line):
                invoices.append(current)
                current = []
                in_invoice = False
    if current:
        invoices.append(current)
    return invoices

# ============================================================
# PRODUCTS + TOTALS PARSER
# ============================================================

def parse_products_and_totals(lines):
    products, totals = [], {}
    tbl_start = None

    for i, l in enumerate(lines):
        if is_tbl_hdr(l):
            tbl_start = i + 1
            break

    if tbl_start is None:
        for i, l in enumerate(lines):
            if PRODUCT_LINE.match(l):
                tbl_start = i
                break

    if tbl_start is None:
        return products, totals

    i = tbl_start
    current = None

    def save():
        nonlocal current
        if current and current.get("qty") is not None:
            products.append(dict(current))

    while i < len(lines):
        l = lines[i].strip()

        if is_subtotal(l):
            save()
            current = None
            v = parse_sar(l)
            totals["subtotal"] = v
            totals["subtotal_value"] = to_float(v)
            i += 1
            continue

        if is_shipping(l):
            if "ﺎًﻧﺎﺠﻣ" in l or "ﺎﻧﺎﺠﻣ" in l:
                totals["shipping"] = "مجانًا"
                totals["shipping_value"] = 0.0
            else:
                v = parse_sar(l)
                totals["shipping"] = v
                totals["shipping_value"] = to_float(v)
            i += 1
            continue

        if is_vat_tot(l) and current is None:
            v = parse_sar(l)
            totals["vat_15"] = v
            totals["vat_15_value"] = to_float(v)
            i += 1
            continue

        if is_ototal(l):
            v = parse_sar(l)
            totals["order_total"] = v
            totals["order_total_value"] = to_float(v)
            i += 1
            continue

        m = PRODUCT_LINE.match(l)
        if m:
            save()
            lt = f"SAR {m.group(1)}"
            up = f"SAR {m.group(2)}"
            current = {
                "name": fix_arabic(m.group(4).strip()),
                "description": "",
                "qty": int(m.group(3)),
                "unit_price": up,
                "unit_price_value": to_float(up),
                "line_total": lt,
                "line_total_value": to_float(lt),
                "vat": None,
                "vat_value": None,
                "price_incl_vat": None,
                "price_incl_vat_value": None,
                "option_type": None,
                "option_value": None,
                "barcode": None,
                "discount_price": None,
            }
            i += 1
            continue

        m = VAT_LINE.match(l)
        if m and current:
            v = f"SAR {m.group(1)}"
            current["vat"] = v
            current["vat_value"] = to_float(v)
            desc = m.group(2).strip()
            if desc:
                current["description"] = fix_arabic(desc)
            i += 1
            continue

        m = PRICE_INCL.match(l)
        if m and current:
            v = f"SAR {m.group(1)}"
            current["price_incl_vat"] = v
            current["price_incl_vat_value"] = to_float(v)
            i += 1
            continue

        m = BEFORE_DISCOUNT.match(l)
        if m and current:
            current["discount_price"] = f"SAR {m.group(1)}"
            i += 1
            continue

        if is_opt_label(l) and current:
            i += 1
            if i < len(lines):
                opt_type, opt_val = parse_option_line(lines[i].strip())
                current["option_type"] = opt_type
                current["option_value"] = opt_val
                i += 1
            continue

        if current and re.match(r"^\d{2,6}-\d+$", l):
            current["barcode"] = l
            i += 1
            continue

        i += 1

    save()
    return products, totals

# ============================================================
# HEADER PARSER
# ============================================================

def parse_header(lines, all_words):
    r = {}

    for w in all_words:
        t = w.get("text", "")
        if re.match(r"^311\d{12}$", t):
            r["vat_number"] = t
        if re.match(r"^\+966\d{9}$", t) and t not in STORE_PHONES:
            r.setdefault("customer_phone", t)

    for i, line in enumerate(lines):
        next_line = lines[i + 1] if i + 1 < len(lines) else ""

        m = re.match(r"^(\d{6,9})\s*:\s*ﺐﻠﻄﻟا", line)
        if not m:
            m = re.match(r"^(\d{6,9})\s*:", line)
        if m and "order_number" not in r:
            r["order_number"] = m.group(1)

        if "vat_number" not in r:
            m = re.search(r"(311\d{12})", line)
            if m:
                r["vat_number"] = m.group(1)

        m = re.match(
            r"((?:Saturday|Sunday|Monday|Tuesday|Wednesday|Thursday|Friday)"
            r"\s+\d{1,2}\s+\w+\s+\d{4})\s*\|",
            line,
        )
        if m and "order_date" not in r:
            dp = m.group(1)
            tm = re.search(r"(\d{2}:\d{2})", line)
            ap = re.search(r"\b(AM|PM)\b", line)
            if tm:
                ampm = ap.group(1) if ap else ""
                if not ampm:
                    ap2 = re.search(r"\b(AM|PM)\b", next_line)
                    ampm = ap2.group(1) if ap2 else ""
                r["order_date"] = f"{dp} | {tm.group(1)} {ampm}".strip()
            else:
                t2 = re.search(r"(\d{2}:\d{2}\s*(?:AM|PM))", next_line)
                if t2:
                    r["order_date"] = f"{dp} | {t2.group(1).strip()}"

        if "ﻎﻠﺒﻤﻟا" in line and "payment_amount" not in r:
            m = re.search(r"SAR\s*([\d,]+(?:\.\d+)?)", line)
            if m:
                r["payment_amount"] = f"SAR {m.group(1)}"

        line_nfkc = nfkc(line)
        if "عفدلا" in line_nfkc and "ةقيرط" in line_nfkc:
            if "payment_method" not in r:
                pm = detect_payment_method(line, next_line)
                if pm:
                    r["payment_method"] = pm
                else:
                    for look in range(1, 4):
                        idx = i + look
                        if idx < len(lines):
                            pm = detect_payment_method(lines[idx])
                            if pm:
                                r["payment_method"] = pm
                                break

    h = next((i for i, l in enumerate(lines) if "ةرﺪﺼﻣ" in l and l.count("ةرﺪﺼﻣ") >= 2), None)
    if h is not None and h + 1 < len(lines):
        name_line = lines[h + 1]
        for kw in ["ﺮﺠﺘﻤﻟا", "ﺔﻴﻏﺎﻓ", "دﻮﻌﻟا", "نيوترﻜﻟﻹا"]:
            idx = name_line.find(kw)
            if idx != -1:
                name_line = name_line[:idx].strip()
        r["customer_name"] = fix_arabic(name_line.strip())
        if h + 2 < len(lines):
            r["customer_country"] = fix_arabic(lines[h + 2].split()[0])
        if h + 3 < len(lines):
            r["customer_city"] = fix_arabic(lines[h + 3].split()[0])

    for line in lines:
        for e in re.findall(r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}\b", line):
            if "vagueoud" not in e:
                r["customer_email"] = e
                break
        if "customer_email" in r:
            break

    return r

# ============================================================
# PARSE SINGLE INVOICE
# ============================================================

def parse_single_invoice(inv_lines, all_words):
    invoice = parse_header(inv_lines, all_words)
    products, totals = parse_products_and_totals(inv_lines)
    return {
        "invoice": invoice,
        "products": products,
        "totals": totals,
        "summary": {
            "total_products": len(products),
            "total_items": sum(p.get("qty") or 0 for p in products),
        },
    }


def extract_all_invoices(pdf_path):
    all_texts, all_words = [], []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            t = normalize_text(page.extract_text() or "")
            all_texts.append(t)
            all_words.extend(page.extract_words() or [])

    all_lines = [l.strip() for l in "\n".join(all_texts).splitlines() if l.strip()]
    blocks = split_invoices(all_lines)

    results = []
    for idx, block in enumerate(blocks, 1):
        parsed = parse_single_invoice(block, all_words)
        parsed["invoice_index"] = idx
        results.append(parsed)
    return results

# ============================================================
# BUILD FLAT TABLE
# ============================================================

INVOICE_COLS = ["order_number", "order_date", "payment_method", "payment_amount", "shipping"]
PRODUCT_COLS = ["product_index", "product_name", "qty", "unit_price", "option_type", "option_value"]


def build_merged_table(all_invoices):
    rows = []
    for inv in all_invoices:
        h     = inv["invoice"]
        tot   = inv["totals"]
        prods = inv["products"]
        order_num = h.get("order_number", f"INV-{inv['invoice_index']}")
        shipping  = tot.get("shipping") or ""

        inv_data = {
            "order_number":   order_num,
            "order_date":     h.get("order_date", ""),
            "payment_method": h.get("payment_method", ""),
            "payment_amount": h.get("payment_amount", ""),
            "shipping":       shipping,
        }

        if not prods:
            row = dict(inv_data)
            row.update({c: None for c in PRODUCT_COLS})
            rows.append(row)
        else:
            for p_idx, p in enumerate(prods, 1):
                row = dict(inv_data)
                row["product_index"] = p_idx
                row["product_name"]  = p.get("name", "")
                row["qty"]           = p.get("qty")
                row["unit_price"]    = p.get("unit_price", "")
                row["option_type"]   = p.get("option_type", "")
                row["option_value"]  = p.get("option_value", "")
                rows.append(row)

    return pd.DataFrame(rows, columns=INVOICE_COLS + PRODUCT_COLS)

# ============================================================
# EXCEL WITH MERGE CELLS
# ============================================================

def style_border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def to_excel_merged(all_invoices, output_path="invoices_output.xlsx"):
    df = build_merged_table(all_invoices)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="الفواتير", index=False)

    wb = load_workbook(output_path)
    ws = wb["الفواتير"]

    HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
    INV_FILL  = PatternFill("solid", fgColor="D6E4F0")
    INV_FILL2 = PatternFill("solid", fgColor="EBF3FB")
    PROD_FILL = PatternFill("solid", fgColor="FFFFFF")
    ALT_FILL  = PatternFill("solid", fgColor="F5F9FD")

    HDR_FONT   = Font(bold=True, color="FFFFFF", size=11)
    BODY_FONT  = Font(size=10)
    MERGE_FONT = Font(bold=True, size=10)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.fill      = HDR_FILL
        cell.font      = HDR_FONT
        cell.alignment = CENTER
        cell.border    = style_border()
    ws.row_dimensions[1].height = 32

    col_widths = {
        "order_number":   18,
        "order_date":     28,
        "payment_method": 22,
        "payment_amount": 18,
        "shipping":       16,
        "product_index":  10,
        "product_name":   42,
        "qty":             8,
        "unit_price":     16,
        "option_type":    18,
        "option_value":   22,
    }
    for col_idx, col_name in enumerate(INVOICE_COLS + PRODUCT_COLS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 16)

    order_rows = defaultdict(list)
    for df_idx, val in enumerate(df["order_number"], start=2):
        order_rows[str(val)].append(df_idx)

    inv_col_n  = len(INVOICE_COLS)
    prod_col_n = len(PRODUCT_COLS)

    for inv_i, (order_num, row_list) in enumerate(order_rows.items()):
        inv_fill = INV_FILL if inv_i % 2 == 0 else INV_FILL2
        r_start  = row_list[0]
        r_end    = row_list[-1]

        for col_idx in range(1, inv_col_n + 1):
            col_letter = get_column_letter(col_idx)
            if r_start != r_end:
                ws.merge_cells(f"{col_letter}{r_start}:{col_letter}{r_end}")
            cell = ws[f"{col_letter}{r_start}"]
            cell.fill      = inv_fill
            cell.font      = MERGE_FONT
            cell.alignment = CENTER
            cell.border    = style_border()

        for row_num in row_list:
            fill = PROD_FILL if row_num % 2 == 0 else ALT_FILL
            for col_idx in range(inv_col_n + 1, inv_col_n + prod_col_n + 1):
                cell = ws[f"{get_column_letter(col_idx)}{row_num}"]
                cell.fill      = fill
                cell.font      = BODY_FONT
                cell.alignment = CENTER if col_idx in [inv_col_n + 1, inv_col_n + 3] else LEFT
                cell.border    = style_border()
            ws.row_dimensions[row_num].height = 28

    ws.freeze_panes           = "A2"
    ws.sheet_view.rightToLeft = True
    wb.save(output_path)
    return output_path, df


# ============================================================
# DEBUG
# ============================================================

def debug_payment(pdf_path):
    all_texts = []
    with pdfplumber.open(pdf_path) as pdf:
        total_pages = len(pdf.pages)
        print(f"إجمالي الصفحات: {total_pages}\n")
        for page_num, page in enumerate(pdf.pages, 1):
            all_texts.append((page_num, normalize_text(page.extract_text() or "")))

    for page_num, text in all_texts:
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        for i, line in enumerate(lines):
            line_n = nfkc(line)
            if "عفدلا" in line_n and "ةقيرط" in line_n:
                prev_l = lines[i - 1] if i > 0 else ""
                next_l = lines[i + 1] if i + 1 < len(lines) else ""
                detected = detect_payment_method(line, next_l)
                print(f"=== صفحة {page_num} ===")
                print(f"  PREV    : {repr(prev_l)}")
                print(f"  THIS    : {repr(line)}")
                print(f"  NFKC    : {repr(line_n)}")
                print(f"  NEXT    : {repr(next_l)}")
                print(f"  DETECTED: {detected}")
                print()
